import os
import re
from openpyxl import Workbook
from datetime import datetime
from app.config import settings
from app.logger import get_logger

logger = get_logger("ExcelService")

def create_deal_report(smart_process_id: int, deals_data: list, sp_name: str = "") -> str:
    """
    Создает Excel файл на основе данных сделок и помещает его в каталог файлов.
    Данные сделок передабются как массив (строки) массивов (значения ячеек в строке массивом).
    Первая строка считается заголовочной.
    Возвращает имя созданного файла.
    """
    try:
        # Выражение для поиска строк с лидирующими нолями
        re_leading_zero = re.compile(r'^0\d+$')
        # Выражение для определение даты по ISO
        re_iso_date = re.compile(r'^\d{4}-\d{2}-\d{2}T')

        filename = normalize_filename(sp_name)
        filename = f"{filename}.xlsx" if filename else f"ticket_register_{smart_process_id}.xlsx"
        filepath = os.path.join(settings.files_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "РеестрБилетов"

        for deal_data in deals_data:
            ws.append(deal_data)
           
            # Переопределяем форматы в ячейках
            for cell in ws[ws.max_row]:
                val = str(cell.value) if cell.value else ""

                # Число с лидирующими нолями
                if re_leading_zero.match(val):
                    cell.number_format = '@'  # Формат Текст
                # Дата в формате ISO
                elif re_iso_date.match(val):
                    try:
                        # Убираем таймзону для Excel и проставляем формат без точного времени
                        cell.value = datetime.fromisoformat(val).replace(tzinfo=None)
                        cell.number_format = 'DD.MM.YYYY' 
                    except ValueError:
                        # Если формат похож, но дата некорректная
                        logger.warning(f"Unable to convert date from ISO string for value {val}")
                        pass


        wb.save(filepath)
        logger.info(f"Excel report created: {filepath}")
        
        return filename

    except Exception as e:
        logger.error(f"Failed to create Excel report: {e}")
        raise e
    
def normalize_filename(name: str) -> str:
    if not name:
        return ""

    # 1. Заменяем один или несколько пробелов (а также табы) на одно подчеркивание
    name = re.sub(r'\s+', '_', name)
    
    # 2. Удаляем все запрещенные для файловых систем символы (Windows/Linux/Mac)
    # Запрещены: < > : " / \ | ? * и служебные (0-31)
    name = re.sub(r'[\\/*?:"<>|\x00-\x1F]', '', name)
    
    # 3. Заменяем несколько подряд идущих подчеркиваний на одно (для красоты)
    name = re.sub(r'_+', '_', name)
    
    # 4. Убираем подчеркивания и точки с краев строки
    name = name.strip('_.')
    
    # 5. Ограничиваем длину (ОС обычно не любят имена длиннее 255 символов)
    # Берем с запасом, чтобы влезло расширение .xlsx
    name = name[:200]
    
    return name